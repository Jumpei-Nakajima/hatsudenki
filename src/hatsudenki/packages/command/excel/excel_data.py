from collections import OrderedDict
from pathlib import Path
from typing import List, Dict, Generator

import openpyxl
from openpyxl.cell import Cell
from openpyxl.styles import NamedStyle, PatternFill, Font
from openpyxl.styles.colors import Color
from openpyxl.worksheet.worksheet import Worksheet


class CellStyle(object):
    NONE = NamedStyle(name='normal')
    BLUE = NamedStyle(name='fill_blue')
    BLUE.fill = PatternFill(fill_type='solid', fgColor='D9E1F2')
    RED = NamedStyle(name='fill_red')
    RED.fill = PatternFill(fill_type='solid', fgColor='FCE4D6')
    GREEN = NamedStyle(name='fill_green')
    GREEN.fill = PatternFill(fill_type='solid', fgColor='CCFFCC')


# HYPERLINK_REGEXP = re.compile('=HYPERLINK\(\"(([#!\.]?[^.]+)+)\", \"(([#!\.]?[^.]+)+)\"\)')


class ExcelData(object):
    def __init__(self, file_path: Path):
        self.file_path = file_path
        self.book: openpyxl.Workbook = None
        self.formula_book: openpyxl.Workbook = None

    @property
    def name(self):
        return self.file_path.name

    def load(self, flg=True):
        if self.book is None:
            self.book = openpyxl.load_workbook(str(self.file_path), data_only=flg)
        if flg:
            self.formula_book = openpyxl.load_workbook(str(self.file_path), data_only=False)

    def create(self):
        # ブック作成
        self.book = openpyxl.Workbook()
        self.book.remove(self.book['Sheet'])

    def get_row_values(self, sheet_name, offset_row, offset_col, width):
        sheet = self.book[sheet_name]
        return [sheet.cell(offset_row + 1, offset_col + i + 1).value for i in range(0, width)]

    def rect_value_iter(self, sheet_name: str, offset_row, offset_col, width):
        sheet: Worksheet = self.book[sheet_name]

        row_idx = 0
        while sheet.cell(row_idx + offset_row + 1, offset_col + 1).value is not None:
            yield self.get_row_values(sheet_name, row_idx + offset_row, offset_col, width)
            row_idx += 1

    def rect_value(self, sheet_name: str, offset_row, offset_col, width, all_str=False):
        ret: List[List[any]] = []
        row = offset_row + 1
        sheet: Worksheet = self.book[sheet_name]
        while sheet.cell(row, offset_col + 1).value is not None:
            if all_str:
                # 全て文字列に変換する
                vals = [str(sheet.cell(row, col + offset_col + 1).value) for col in range(0, width)]
            else:
                vals = [sheet.cell(row, col + offset_col + 1).value for col in range(0, width)]
            ret.append(vals)
            row += 1
        return ret

    def rect_value_to_dict(self, sheet_name, offset_row, offset_col, width):
        ret = []
        r = offset_row + 1
        sheet = self.book[sheet_name]

        # 一行目をヘッダとする
        header = [sheet.cell(r, x + offset_col + 1).value for x in range(width)]

        r += 1

        while sheet.cell(r, 1).value is not None:
            vals = {header[x]: sheet.cell(r, x + offset_col + 1).value for x in range(0, width)}
            ret.append(vals)
            r += 1
        return ret

    def get_sheet(self, sheet_name: str, formula: bool = False) -> Worksheet:
        if not formula or self.formula_book is None:
            return self.book[sheet_name]

        return self.formula_book[sheet_name]

    def iter_row(self, sheet_name: str) -> Generator[List[Cell], None, None]:
        sheet = self.get_sheet(sheet_name)
        return sheet.rows

    def iter_row_dict(self, sheet_name: str):
        sheet = self.get_sheet(sheet_name, formula=True)
        iter = sheet.rows
        h = next(iter, None)
        if h is None:
            return []

        key_names = []
        for cell in h:
            c = cell.value
            if c is None:
                key_names.append('column_' + str(len(key_names)))
            else:

                if c.startswith('=HYPERLINK'):
                    v1, v2 = c.split(',')
                    c = v2.strip()[1:-2]

                # m = HYPERLINK_REGEXP.match(c)
                # if m is not None:
                #     c = m.group(3)

                key_names.append(c)

        for cells in iter:
            v = [cell.value for cell in cells]
            if len(v) is 0 or v[0] is None:
                break
            l = zip(key_names, v)
            yield OrderedDict(l)

        # return (OrderedDict(zip(key_names, [cell.value for cell in cells])) for cells in iter)

    def max_column(self, sheet_name):
        sheet = self.book[sheet_name]
        return sheet.max_column

    def max_row(self, sheet_name):
        sheet = self.book[sheet_name]

        return sheet.max_row

    def add_named_style(self, style):
        if style.name in self.book.named_styles:
            return
        self.book.add_named_style(style)

    def write_row(self, sheet_name, row, col, data, style=CellStyle.NONE, header_links: List[str] = None,
                  column_size: Dict[str, int] = None):
        header_links: List[str] = header_links or []
        column_size: Dict[str, int] = column_size or {}
        self.add_named_style(style)
        i = 0
        sheet = self.get_sheet(sheet_name)
        for cell in (sheet.cell(row + 1, col + 1 + idx) for idx in range(len(data))):
            cell.value = data[i]
            cell.style = style.name
            column_size[cell.column] = max((column_size.get(cell.column, 0), len(str(cell.value))))

            if header_links is not None and len(header_links) > i and len(header_links[i]) > 0:
                cell.value = f'=HYPERLINK("{header_links[i]}", "{cell.value}")'
                cell.font = Font(name="Meiryo UI", size=10, underline="single",
                                 color=Color(rgb=None, indexed=None, auto=None, theme=10, tint=0.0, type="theme"))

            i += 1

    def insert_row(self, sheet_name, row, col, data, style=CellStyle.NONE):
        sheet = self.get_sheet(sheet_name)
        sheet.insert_rows(row + 1, 1)
        self.write_row(sheet_name, row, col, data, style)

    def save_file(self):
        if len(self.book.worksheets) is 0:
            return
        self.file_path.parent.mkdir(parents=True, exist_ok=True)
        self.book.save(str(self.file_path))

    @classmethod
    def init_column_size(cls, st: Worksheet, column_size: Dict[str, int]):
        for col, value in column_size.items():
            st.column_dimensions[col].width = max(value * 2, 10)

    def get_header_values(self, sheet_name):
        book = self.formula_book or self.book

        first_line = next(book[sheet_name].rows)
        ret = {}
        for cell in first_line:
            c = cell.value
            if c is None:
                continue
            c = str(c)
            if c.startswith('=HYPERLINK'):
                v1, v2 = c.split(',')
                c = v2.strip()[1:-2]

            ret[c] = cell.coordinate
        return ret
