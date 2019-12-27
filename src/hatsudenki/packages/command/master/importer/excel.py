from pathlib import Path
from typing import List, Dict

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from hatsudenki.packages.command.stdout.output import ToolOutput


class ImportMasterExcelSheet(object):
    def __init__(self, book: 'ImportMasterExcel', sheet):
        self.book = book
        self.sheet: Worksheet = sheet

    def add_table(self, table_data: List[dict]):
        ToolOutput.anchor(f'{self.book.path.name} {self.sheet.title}にレストア')
        r = self.sheet.rows
        headers = next(r)
        idxes = [h.value for h in headers]

        self.sheet.delete_rows(2, self.sheet.max_row - 1)

        for d in table_data:
            dd = []
            for idx in idxes:
                if idx == 'TAG':
                    # TAGは出力されないので強制的に0入れる
                    dd.append(0)
                    continue
                dd.append(d.get(idx, None))
            self.sheet.append(dd)
        ToolOutput.pop('OK')


class ImportMasterExcel(object):
    def __init__(self, path: Path):
        self.path = path
        self.book: openpyxl.Workbook = None
        self.sheets: Dict[str, ImportMasterExcelSheet] = {}
        self.load()

    def load(self):
        if self.book is None:
            self.book = openpyxl.load_workbook(str(self.path), data_only=False)
            self.sheets = {sn: ImportMasterExcelSheet(self, self.book[sn]) for sn in self.book.sheetnames}

    def save(self):
        self.book.save(self.path)
