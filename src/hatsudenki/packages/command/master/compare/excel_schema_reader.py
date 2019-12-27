from __future__ import annotations

import ntpath
import re
import unicodedata
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from hatsudenki.packages.command.master.compare.schema import Schema


class ExcelSchemaReader(object):
    def __init__(self, base_path: Path):
        self.base_path = base_path
        self._data = {}

    def setup(self) -> ExcelSchemaReader:
        print(f'loading...xlsx files...')
        return self

    def schema(self) -> Schema:
        return Schema([self._read_group(p) for p in self.base_path.glob('*.xlsx') if not p.name.startswith('~$')])

    def _read_group(self, path) -> Schema.Group:
        grop_name = ntpath.basename(unicodedata.normalize('NFC', str(path))).replace('.xlsx', '')
        book = openpyxl.load_workbook(path)
        print(f'    loading...{grop_name}')
        return Schema.Group(grop_name, [t for t in (self._read_table(s) for s in book.worksheets) if t is not None])

    def _read_table(self, sheet: Worksheet) -> Optional[Schema.Group.Table]:
        if self._is_exclude_sheet(sheet):
            print(f'        skipping...{sheet.title}')
            return None
        print(f'        loading...{sheet.title}')
        table_name = sheet.title
        row = (c for c in (sheet.cell(1, i) for i in range(1, sheet.max_column)) if not self._is_exclude_cell(c))
        return Schema.Group.Table(table_name, [col for col in (self._read_column(c) for c in row) if col is not None])

    @staticmethod
    def _is_exclude_sheet(s: Worksheet):
        return s.title.startswith('★')

    def _read_column(self, cell: Cell) -> Optional[Schema.Group.Table.Column]:
        HYPERLINK_LEFT_REGEXP = re.compile('=HYPERLINK\(\"(([#!\.]?[^.]+)+)\", \"')
        cell_value = cell.value
        match_result = HYPERLINK_LEFT_REGEXP.match(cell_value)
        if match_result:
            delete_str = match_result.group()
            cell_value = cell_value.replace(delete_str, "")
            cell_value = cell_value.replace('")', "")
        return Schema.Group.Table.Column(cell_value)

    valid_colors = ['00CCFFCC', 'FFFCE4D6', '00FCE4D6']

    def _is_exclude_cell(self, c: Cell):
        if c.value is None:
            return True
        if c.fill.fgColor.value not in self.valid_colors:
            print(f"            skipping column {c.value} : {c.fill.fgColor.value}")
        return c.fill.fgColor.value not in self.valid_colors
