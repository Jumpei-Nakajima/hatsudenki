from pathlib import Path

from openpyxl.styles import Font, Color

from hatsudenki.packages.command.excel.excel_data import ExcelData, CellStyle
from hatsudenki.packages.command.master.enum.loader import EnumLoader
from hatsudenki.packages.command.master.loader import MasterTableLoader


class EnumExcelBook(ExcelData):

    def __init__(self, file_path: Path, enum_loader: EnumLoader, table_loader: MasterTableLoader):
        super().__init__(file_path)
        self.enum_loader = enum_loader
        self.table_loader = table_loader
        self.create()
        self._generate_sheets()

    def _generate_sheets(self):
        for k, data in self.enum_loader.iter():
            sheet_name = data.excel_sheet_name
            if sheet_name in self.book.sheetnames:
                self.book.remove(self.book[sheet_name])
            sheet = self.book.create_sheet(sheet_name)
            sheet.protection.enable()
            sheet.sheet_properties.tabColor = 'ff0000'

            if data.is_selector:
                self.write_row(sheet_name, 0, 0, ['id', 'name', 'ref_name', 'display_name', 'value', 'comment'],
                               CellStyle.RED)
                for idx, v in enumerate(data.values):
                    row = v.excel_row
                    val = row[4]
                    if val.startswith('enum_'):
                        target_enum = self.enum_loader.get_by_name(val)
                        link_label = f'{target_enum.excel_name}.xlsx#{target_enum.excel_sheet_name}!A1'
                        link_value = f'【参照】{target_enum.excel_name} [{target_enum.excel_sheet_name}]'
                    else:
                        target_table = self.table_loader.get_by_table_name(val)
                        if target_table is None:
                            link_label = ''
                            link_value = '【参照先定義なし】'
                            print('=================================================')
                            print(f'参照先なし:{val}')
                        else:
                            link_label = f'../{target_table.excel_name}.xlsx#{target_table.excel_sheet_name}!A1'
                            link_value = f'【参照】{target_table.excel_name} [{target_table.excel_sheet_name}]'
                    self.write_row(sheet_name, idx + 1, 0, row)
                    c = self.book[sheet_name].cell(idx + 2, 5)
                    c.value = link_value
                    c.hyperlink = link_label
                    c.font = Font(name="Meiryo UI", size=10, underline="single",
                                  color=Color(rgb=None, indexed=None, auto=None, theme=10, tint=0.0, type="theme"))

            else:
                self.write_row(sheet_name, 0, 0, ['id', 'name', 'ref_name', 'display_name', 'comment'],
                               CellStyle.RED)
                for idx, v in enumerate(data.values):
                    self.write_row(sheet_name, idx + 1, 0, v.excel_row)
